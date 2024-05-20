import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl

cap = cv2.VideoCapture(1)

entrada = set()
salida = set()

def infhora():
    inf = datetime.now()
    # extraer fecha y hora
    fecha = inf.strftime('%Y:%m:%d')
    hora = inf.strftime('%H:%M:%S')
    return hora, fecha

# Inicializar el workbook y crear las hojas fuera del bucle while
wb = xl.Workbook()
hojam = wb.create_sheet("Entrada")
hojan = wb.create_sheet("Salida")

# Eliminar la hoja predeterminada creada por Workbook
del wb['Sheet']

# Principal
while True:
    # leer frames
    ret, frame = cap.read()

    # interfaz y texto
    cv2.putText(frame, 'Locate the QR code', (160, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    # rectángulo
    cv2.rectangle(frame, (170, 100), (470, 400), (0, 255, 0), 2)

    # extraer hora y fecha
    hora, fecha = infhora()
    diasem = datetime.today().weekday()

    print(diasem)
    # año, mes, día
    a, me, d = fecha[0:4], fecha[5:7], fecha[8:10]
    # hora, min, seg
    h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

    # crear nombre de archivo Excel
    nomar = f"{a}-{me}-{d}"
    texth = f"{h}-{m}-{s}"
    print(texth)

    for codes in decode(frame):
        # decodificar la información
        info = codes.data.decode('utf-8')

        # tipo de persona
        tipo = info[0:2]
        tipo = int(tipo)
        letr = chr(tipo)

        # número
        num = info[2:]

        # Extraer coordenadas
        pts = np.array([codes.polygon], np.int32)
        xi, yi = codes.rect.left, codes.rect.top

        # id completo
        codigo = letr + num

        # días de la semana
        if 0 <= diasem <= 4:
            # dividir horas del día
            # mañana
            if 7 <= h <= 10:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                # guardar id
                if codigo not in entrada:
                    entrada.add(codigo)

                    # guardar en la hoja de entrada
                    hojam.append([codigo,hora])
                    wb.save(nomar + '.xlsx')
                    
                else:
                    cv2.putText(frame, f'EL ID {codigo}', (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    cv2.putText(frame, 'Fue registrado', (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        if 17 < h <= 19:
            cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
            if codigo not in salida:
                salida.add(codigo)
                cv2.putText(frame, letr + '0' + str(num), (xi - 15, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                
                # guardar en la hoja de salida
                hojan.append([codigo,hora])
                wb.save(nomar + '.xlsx')
            else:
                cv2.putText(frame, f'El Roster {codigo}', (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                cv2.putText(frame, 'Fue registrado', (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

    cv2.imshow("QR READER", frame)
    t = cv2.waitKey(5)
    if t == 27:
        break

cv2.destroyAllWindows()
cap.release()
